import pandas as pd
import streamlit as st
import matplotlib.pyplot as plt
import openpyxl
import streamlit as st
import pandas as pd
import datetime
import streamlit as st
import pandas as pd
import os

st.set_page_config(
    page_title="Aplikasi Monitoring Samas",
    page_icon=":chart_with_upwards_trend:",
    layout="wide",
)

def createInvoice(nomor, terima_dari, pekerjaan, jenis_muatan, harga, tanggal, nomor_volume):
    # Specify the path to the Excel file
    if type(nomor_volume) == str:
        file_path = 'C:/Users/salma/pyworkspace/KW-PERMATA BANK.xlsx'
    else:
        file_path = 'C:/Users/salma/pyworkspace/KW-LTMPLB-2023 - Contoh.xlsx'
    workbook = openpyxl.load_workbook(file_path)
    # Select the active sheet
    sheet = workbook.active

    # Rewrite nomor
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value == "119/SAMAS-KW/LTM/VIII/2023":
                cell.value = nomor

    # Rewrite terima_dari
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value == "PT. ALAM MULTI MEGA":
                cell.value = terima_dari
                
    # Rewrite pekerjaan
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value == "Biaya Jasa Angkutan Darat Dari PT SAP ke Jetty LKS ":
                cell.value = pekerjaan
    
    # Rewrite jenis_muatan
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value == "Cangkang Sawit":
                cell.value = jenis_muatan
    
    # Rewrite nomor spk atau volume
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value == 207970:
                cell.value = nomor_volume
            elif cell.value == "Nomor SPK : 065/SPK-PM/CRES/II/2023":
                cell.value = "Nomor SPK : " + nomor_volume
    
    # Rewrite Harga Total
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value == 20797000:
                cell.value = harga
    
    ppn = (harga * 11) / 100
    # Rewrite PPN
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value == "=N14*11%":
                cell.value = ppn
    
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value == "=S1":
                cell.value = jenis_muatan
    
    bersih = harga + ppn 
    # Rewrite bersih
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value == "=N14+N15":
                cell.value = bersih
    
    # Rewrite bersih yang besar
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value == "=N16":
                cell.value = bersih
    
    # Rewrite tanggal
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value == "Palembang, 11 Agustus 2023":
                cell.value = "Palembang, " + tanggal
    
    # Save the workbook
    sheet.sheet_view.showGridLines = False
    workbook.save('../invoice.xlsx')
    return True
 
st.header("Buat Invoice")

# Create a list of unique values in the column 'Project'
nomor = st.text_input(label='Nomor Invoice: ', placeholder="Nomor", value='')
terima_dari = st.text_input(label='Telah Terima Dari: ', placeholder="Nama Perusahaan", value='')
pekerjaan = st.text_input(label='Pekerjaan: ', placeholder="Keterangan Pekerjaan", value='')
jenis_muatan = st.selectbox('Jenis Muatan: ', ("-","Cangkang Sawit", "Cocopeat", "Kopra", "Jagung", "Kelapa", "Sekam", "Pupuk", "Bibit"), placeholder="Pilih Jenis Muatan")
if jenis_muatan == "-":
    harga = st.number_input(label="Harga (Rp): ", placeholder="Rp ", value=0)
    nomor_volume = st.text_input(label="Nomor SPK: ", placeholder="Nomor SPK", value="")
    volume = False
else:
    hcol1, hcol2 = st.columns(2)
    with hcol1:
        nomor_volume = st.number_input(label='Volume (Kg): ', placeholder="Kg", value=0)
    with hcol2: 
        harga_barang = st.number_input(label='Harga Per Kilo (Rp): ', placeholder="Rp", value=0)
    harga = nomor_volume * harga_barang
tanggal = st.date_input('Tanggal: ', datetime.datetime.now(datetime.timezone(datetime.timedelta(hours=7))))
tanggal = tanggal.strftime("%d %B %Y")

filename = "Invoice {nomor}.xlsx".format(nomor=nomor)
invoice = False
col1, col2 = st.columns(2)
with col1:
    if st.button('Buat Invoice'):
        invoice = createInvoice(nomor, terima_dari, pekerjaan, jenis_muatan, harga, tanggal, nomor_volume)
with col2:
    with open("../invoice.xlsx", "rb") as f:
        st.download_button(label = 'Download Invoice',
                            data=f.read(),
                            file_name=filename,
                            disabled=False if invoice else True,
                            mime='application/vnd.ms-excel',)
st.header("Buat Laporan Otomatis")
uploaded_file = st.file_uploader('Upload File Excel')
 
if uploaded_file:
    df = pd.read_csv(uploaded_file)
    st.dataframe(df)

st.button('Buat Laporan')

st.header("Dashboard Cash In - Out")
# Load data from Excel file
laporan_cash_in_out = pd.ExcelFile("C:/Users/salma/pyworkspace/Datasets/Contoh Laporan Cash In -Out Januari 23.xlsx")
cash_df = pd.read_excel(laporan_cash_in_out).fillna(0)
min_date = cash_df.Tanggal.min()
max_date = cash_df.Tanggal.max()
with st.sidebar:
    date_range = st.date_input(
        label="Select Date Range",
        min_value=min_date,
        max_value=max_date,
        value=[min_date, max_date],
        format="YYYY.MM.DD",
    )
    if len(date_range) == 1:
        main_df = cash_df
    else:
        start_date, end_date = date_range
        main_df = cash_df[(cash_df.Tanggal >= str(start_date)) & (cash_df.Tanggal <= str(end_date))]

ss_01 = main_df.loc[(main_df["Project"] == "SS-01")]
ss_02 = main_df.loc[(main_df["Project"] == "SS-02")]
ss_03 = main_df.loc[(main_df["Project"] == "SS-03")]
ss_05 = main_df.loc[(main_df["Project"] == "SS-05")]
ss_06 = main_df.loc[(main_df["Project"] == "SS-06")]
ss_09 = main_df.loc[(main_df["Project"] == "SS-09")]
tab1, tab2, tab3, tab4, tab5, tab6 = st.tabs(["SS-09", "SS-02", "SS-03", "SS-05", "SS-06", "SS-01"])

with tab6:
    # Create bar plot
    fig, ax = plt.subplots()
    ax.pie(ss_01[["Kredit", "Debit"]].sum(), labels=["Kredit", "Debit"], autopct="%1.1f%%", startangle=90)

    # Set plot title and labels
    ax.set_title("SS - 09 Cash In - Out")

    # Display the plot in Streamlit
    st.pyplot(fig)
    st.write(ss_01.tail(5))

with tab2:
    # Create bar plot
    fig, ax = plt.subplots()
    ax.pie(ss_02[["Kredit", "Debit"]].sum(), labels=["Kredit", "Debit"], autopct="%1.1f%%", startangle=90)
    
    # Set plot title and labels
    ax.set_title("SS - 02 Cash In - Out")
    
    # Display the Plot in Streamlit
    st.pyplot(fig)
    st.write(ss_02.tail(5))
    
with tab3:
    fig, ax = plt.subplots()
    ax.pie(ss_03[["Kredit", "Debit"]].sum(), labels=["Kredit", "Debit"], autopct="%1.1f%%", startangle=90)

    # Set plot title and labels
    ax.set_title("SS - 03 Cash In - Out")

    # Display the plot in Streamlit
    st.pyplot(fig)
    st.write(ss_03.tail(5))

with tab4:
    fig, ax = plt.subplots()
    ax.pie(ss_05[["Kredit", "Debit"]].sum(), labels=["Kredit", "Debit"], autopct="%1.1f%%", startangle=90)

    # Set plot title and labels
    ax.set_title("SS - 05 Cash In - Out")

    # Display the plot in Streamlit
    st.pyplot(fig)
    st.write(ss_05.tail(5))
    
with tab5:
    fig, ax = plt.subplots()
    ax.pie(ss_06[["Kredit", "Debit"]].sum(), labels=["Kredit", "Debit"], autopct="%1.1f%%", startangle=90)

    # Set plot title and labels
    ax.set_title("SS - 06 Cash In - Out")

    # Display the plot in Streamlit
    st.pyplot(fig)
    st.write(ss_06.tail(5))
    
with tab1:
    fig, ax = plt.subplots()
    ax.pie(ss_09[["Kredit", "Debit"]].sum(), labels=["Kredit", "Debit"], autopct="%1.1f%%", startangle=90)

    # Set plot title and labels
    ax.set_title("SS - 01 Cash In - Out")

    # Display the plot in Streamlit
    st.pyplot(fig)
    st.write(ss_09.tail(5))
    
